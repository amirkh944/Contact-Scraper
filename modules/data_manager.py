import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
from datetime import datetime
import os

class DataManager:
    """مدیریت کننده داده ها"""
    
    def __init__(self, output_dir: str = 'data/results'):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_to_excel(self, data: List[Dict], filename: str = None) -> str:
        """
        صادر کردن داده ها به فایل Excel
        
        Args:
            data: داده های مورد نیاز برای صادر
            filename: نام فایل
            
        Returns:
            مسیر فایل صادر شده
        """
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "نتایج"
        
        # Headers
        headers = ['وبسایت', 'دامنه', 'لینک', 'ایمیل', 'تماس']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Add data
        for item in data:
            ws.append([
                item.get('website_name', ''),
                item.get('domain', ''),
                item.get('website_url', ''),
                ', '.join(item.get('emails', [])),
                ', '.join(item.get('phones', [])),
            ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        
        wb.save(filepath)
        return filepath
    
    def export_to_csv(self, data: List[Dict], filename: str = None) -> str:
        """
        صادر کردن داده ها به فایل CSV
        
        Args:
            data: داده های مورد نیاز برای صادر
            filename: نام فایل
            
        Returns:
            مسیر فایل صادر شده
        """
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['website_name', 'domain', 'website_url', 'emails', 'phones']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for item in data:
                writer.writerow({
                    'website_name': item.get('website_name', ''),
                    'domain': item.get('domain', ''),
                    'website_url': item.get('website_url', ''),
                    'emails': ', '.join(item.get('emails', [])),
                    'phones': ', '.join(item.get('phones', [])),
                })
        
        return filepath
    
    def export_to_json(self, data: List[Dict], filename: str = None) -> str:
        """
        صادر کردن داده ها به فایل JSON
        
        Args:
            data: داده های مورد نیاز برای صادر
            filename: نام فایل
            
        Returns:
            مسیر فایل صادر شده
        """
        if not filename:
            filename = f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        filepath = os.path.join(self.output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return filepath